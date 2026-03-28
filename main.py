import ast
import logging
import math
from itertools import product

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BOWSTYLE_CODES = ["R", "B", "C", "L"]
EXPERIENCE_CODES = ["E", "N"]
GENDER_CODES = ["O", "W"]

BOWSTYLE_NAMES = {"R": "Recurve", "B": "Barebow", "C": "Compound", "L": "Longbow"}
GENDER_NAMES = {"O": "Open", "W": "Women"}

URL_YEAR = 2026

# Columns excluded when building result output (keep all score columns)
META_COLS = frozenset(
    {"Pos.", "Country", "class (with Exp)", "class", "club code", "placelimit"}
)

# Thick red border used to mark the qualification cutoff
CUTOFF_BORDER_SIDE = Side(style="thick", color="FF0000")


def _flatten_columns(columns) -> list:
    """Return clean one-level column names from plain or MultiIndex headers."""
    if isinstance(columns, pd.MultiIndex):
        return [str(c).strip() for c in columns.get_level_values(-1)]
    return [str(c).strip() for c in columns]


def _make_unique_columns(columns: list) -> list:
    """Ensure column labels are unique while preserving original order."""
    counts: dict[str, int] = {}
    unique = []
    for col in columns:
        key = col if col else "column"
        n = counts.get(key, 0)
        counts[key] = n + 1
        unique.append(key if n == 0 else f"{key}_{n + 1}")
    return unique


def _clean_scraped_table(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize table headers and keep only true archer result rows."""
    df = df.dropna(how="all").copy()
    df.columns = _flatten_columns(df.columns)

    # Remove fully empty columns created by decorative headers/footers.
    df = df.loc[:, ~df.columns.astype(str).str.lower().str.startswith("unnamed:")]
    df.columns = _make_unique_columns([str(c).strip() for c in df.columns])

    # Strip whitespace/non-breaking spaces in object columns.
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
            .replace({"": pd.NA, "nan": pd.NA})
        )

    # Keep only the numeric score before '/' in end-by-end columns.
    for col in ("20y-1", "20y-2"):
        if col in df.columns:
            extracted = df[col].astype(str).str.extract(r"^\s*(\d+)")[0]
            df[col] = pd.to_numeric(extracted, errors="coerce")

    if "Tot." in df.columns:
        df["Tot."] = pd.to_numeric(df["Tot."], errors="coerce")
        df = df[df["Tot."].notna()].copy()

    return df


# ---------------------------------------------------------------------------
# 1. Initialisation
# ---------------------------------------------------------------------------
def parse_input_params(filepath: str) -> dict:
    """Read key = value pairs from inputParams.txt via ast.literal_eval."""
    params = {}
    with open(filepath, "r") as fh:
        for line in fh:
            line = line.strip()
            if "=" in line:
                key, _, value = line.partition("=")
                params[key.strip()] = ast.literal_eval(value.strip())
    return params


# ---------------------------------------------------------------------------
# 2. Data scraping
# ---------------------------------------------------------------------------
def scrape_all_data(tournament_ids: list) -> pd.DataFrame:
    """
    Iterate over all tournament IDs and code permutations, append HTML tables
    into a single DataFrame.  Missing categories are silently skipped.
    """
    all_frames = []

    for tournament_id in tournament_ids:
        found = 0
        for bowstyle, experience, gender in product(
            BOWSTYLE_CODES, EXPERIENCE_CODES, GENDER_CODES
        ):
            code = f"IQ{bowstyle}{experience}{gender}"
            url = (
                f"https://www.ianseo.net/TourData/{URL_YEAR}"
                f"/{tournament_id}/{code}.php"
            )

            try:
                tables = pd.read_html(url, flavor="lxml")
                if not tables:
                    continue
                df = _clean_scraped_table(tables[0])
                if df.empty:
                    continue
                df["class (with Exp)"] = f"{bowstyle}{experience}{gender}"
                all_frames.append(df)
                found += 1
            except Exception:
                pass  # silently skip missing / empty categories

        logger.info(
            f"Scraping complete for tournament {tournament_id} "
            f"({found} categories found)"
        )

    if not all_frames:
        raise ValueError("No data scraped from any tournament or category.")

    return pd.concat(all_frames, ignore_index=True)


# ---------------------------------------------------------------------------
# 3. Data parsing
# ---------------------------------------------------------------------------
def parse_data(df: pd.DataFrame, club_code: int, finals_capacity: int) -> dict:
    """
    Derive all computed columns and return a results dictionary containing
    the enriched DataFrame plus summary structures.
    """
    # 3a: class column (remove middle experience letter from 3-letter code)
    df["class"] = df["class (with Exp)"].str[0] + df["class (with Exp)"].str[2]

    # 3b: split Country into club code / club name
    split = df["Country"].str.split(" - ", n=1, expand=True)
    df["club code"] = split[0].str.strip()
    df["club name"] = split[1].str.strip() if 1 in split.columns else ""

    # 3c: ensure Tot. is numeric
    df["Tot."] = pd.to_numeric(df["Tot."], errors="coerce")
    for col in ("Hits", "Golds"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 3d: national rank per class (ties: Tot., then Hits, then Golds)
    sort_cols = ["class", "Tot."]
    ascending = [True, False]
    if "Hits" in df.columns:
        sort_cols.append("Hits")
        ascending.append(False)
    if "Golds" in df.columns:
        sort_cols.append("Golds")
        ascending.append(False)

    df = df.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)
    df["national rank"] = df.groupby("class").cumcount() + 1

    # 3e: entry counts per class and total
    entries_per_class = df.groupby("class").size()
    total_entries = int(entries_per_class.sum())

    # 3f: placelimit per class with iterative minimum-8 allocation
    placelimits = _allocate_placelimits(entries_per_class, finals_capacity, min_places=8)

    # 3g: qualifying score per class (score of archer at rank == placelimit)
    qual_scores: dict = {}
    for cls, limit in placelimits.items():
        class_df = df[df["class"] == cls].sort_values("national rank")
        if len(class_df) >= limit:
            qual_scores[cls] = class_df.iloc[limit - 1]["Tot."]
        else:
            qual_scores[cls] = class_df.iloc[-1]["Tot."]

    # 3h: attach placelimit to df for safety margin calculation
    df["placelimit"] = df["class"].map(placelimits)

    # 3i: club-filtered subset with safety margin
    club_code_str = str(club_code)
    club_df = df[df["club code"] == club_code_str].copy()
    club_df["safety margin"] = (
        (club_df["placelimit"] - club_df["national rank"]) / club_df["placelimit"]
    )
    club_df["safety margin"] = (club_df["safety margin"] * 100).round(2)
    club_df["qualification score"] = club_df["class"].map(qual_scores)
    club_df["qualification rank"] = club_df["class"].map(placelimits)

    return {
        "df": df,
        "entries_per_class": entries_per_class,
        "placelimits": placelimits,
        "qual_scores": qual_scores,
        "club_df": club_df,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _result_cols(df: pd.DataFrame) -> list:
    """
    Ordered output columns for result sheets:
    national rank, Athlete, club name, then all remaining score columns.
    """
    fixed = ["national rank", "Athlete", "club name"]
    score_cols = [
        c for c in df.columns
        if c not in META_COLS and c not in fixed and c != "national rank"
    ]
    return fixed + score_cols


def _allocate_placelimits(
    entries_per_class: pd.Series,
    finals_capacity: int,
    min_places: int = 8,
) -> dict:
    """
    Allocate finals places while enforcing the class minimum.
    Classes that fall below the minimum under proportional allocation are fixed
    first; remaining capacity is distributed proportionally to other classes.
    """
    counts = {cls: int(count) for cls, count in entries_per_class.items()}
    remaining = set(counts.keys())
    fixed: dict = {}
    remaining_capacity = int(finals_capacity)
    remaining_entries = int(sum(counts.values()))

    # Iteratively lock classes that require the minimum places.
    while remaining and remaining_capacity > 0 and remaining_entries > 0:
        newly_fixed = []
        for cls in sorted(remaining):
            proportional = counts[cls] * remaining_capacity / remaining_entries
            if math.floor(proportional) < min_places:
                newly_fixed.append(cls)

        if not newly_fixed:
            break

        for cls in newly_fixed:
            fixed[cls] = min_places
            remaining.remove(cls)
            remaining_capacity -= min_places
            remaining_entries -= counts[cls]

    placelimits = dict(fixed)

    if remaining and remaining_capacity > 0 and remaining_entries > 0:
        exact = {
            cls: counts[cls] * remaining_capacity / remaining_entries
            for cls in remaining
        }
        base = {cls: math.floor(value) for cls, value in exact.items()}

        for cls, value in base.items():
            placelimits[cls] = value

        assigned = sum(placelimits.values())
        leftovers = max(0, finals_capacity - assigned)

        if leftovers > 0:
            order = sorted(
                remaining,
                key=lambda cls: (exact[cls] - base[cls], counts[cls]),
                reverse=True,
            )
            for idx in range(leftovers):
                target = order[idx % len(order)]
                placelimits[target] += 1

    for cls in counts:
        placelimits.setdefault(cls, min_places)

    return placelimits


def _title_case_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Capitalize each word in output column headings."""
    out = df.copy()
    renamed_cols = []
    for col in out.columns:
        col_str = str(col)
        if col_str == "Tot.":
            renamed_cols.append("Score")
        elif col_str == "qualification score":
            renamed_cols.append("Score Required To Qualify")
        elif col_str == "qualification rank":
            renamed_cols.append("Rank Required To Qualify")
        else:
            renamed_cols.append(col_str.replace("_", " ").title())
    out.columns = renamed_cols
    if out.index.name:
        out.index.name = str(out.index.name).replace("_", " ").title()
    return out


def _pivot_table(
    entries_per_class: pd.Series,
    placelimits: dict,
    qual_scores: dict,
    value_source: str,
) -> pd.DataFrame:
    """Build a bowstyle x gender summary table."""
    rows = [BOWSTYLE_NAMES[b] for b in BOWSTYLE_CODES]
    cols = [GENDER_NAMES[g] for g in GENDER_CODES]
    table = pd.DataFrame("-", index=rows, columns=cols, dtype=object)

    for b in BOWSTYLE_CODES:
        for g in GENDER_CODES:
            cls = b + g
            if value_source == "entries":
                raw = entries_per_class.get(cls, 0)
                val = int(raw) if raw else "-"
            elif value_source == "placelimit":
                val = placelimits.get(cls, "-")
            else:  # qual_score
                val = qual_scores.get(cls, "-")
            table.loc[BOWSTYLE_NAMES[b], GENDER_NAMES[g]] = val

    table.index.name = "Bowstyle \\ Gender"
    return table


def _apply_cutoff_border(ws, limit: int):
    """
    Draw a thick red top border on the first row after the placelimit cutoff.
    Row 1 is the header; first data row is row 2.
    """
    first_non_qual = limit + 2  # +1 for header, +1 to get next row
    if first_non_qual > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=first_non_qual, column=col)
        existing = cell.border
        cell.border = Border(
            top=CUTOFF_BORDER_SIDE,
            left=existing.left,
            right=existing.right,
            bottom=existing.bottom,
        )


def _auto_width(ws):
    """Set column widths to fit content."""
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length, 45)


# ---------------------------------------------------------------------------
# 4. Output
# ---------------------------------------------------------------------------
def write_output(parsed: dict, output_path: str):
    df = parsed["df"]
    placelimits = parsed["placelimits"]
    qual_scores = parsed["qual_scores"]
    entries_per_class = parsed["entries_per_class"]
    club_df = parsed["club_df"]

    result_columns = _result_cols(df)

    # Summary pivot tables
    entries_table = _pivot_table(entries_per_class, placelimits, qual_scores, "entries")
    rank_table = _pivot_table(entries_per_class, placelimits, qual_scores, "placelimit")
    score_table = _pivot_table(entries_per_class, placelimits, qual_scores, "qual_score")

    # sheet_name -> placelimit for post-processing borders
    cutoff_info: dict = {}

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Per-class result sheets
        for cls in sorted(df["class"].unique()):
            sheet_name = cls

            class_df = (
                df[df["class"] == cls]
                .sort_values("national rank")
                .reset_index(drop=True)
            )
            out_cols = [c for c in result_columns if c in class_df.columns]
            class_out = _title_case_headers(class_df[out_cols])
            class_out.to_excel(writer, sheet_name=sheet_name, index=False)
            cutoff_info[sheet_name] = placelimits.get(cls, 0)

        # Summary sheets
        _title_case_headers(entries_table).to_excel(
            writer,
            sheet_name="Qualification_Entry_Numbers",
        )
        _title_case_headers(rank_table).to_excel(writer, sheet_name="Rank_to_Qualify")
        _title_case_headers(score_table).to_excel(writer, sheet_name="Score_to_Qualify")

        # Club results sheet
        club_result_cols = [c for c in result_columns if c in club_df.columns]
        extra_cols = [
            c
            for c in [
                "class",
                "qualification score",
                "qualification rank",
                "safety margin",
            ]
            if c in club_df.columns
        ]
        all_club_cols = list(dict.fromkeys(club_result_cols + extra_cols))
        club_out = (
            club_df[all_club_cols]
            .sort_values("safety margin", ascending=False)
            .reset_index(drop=True)
        )

        _title_case_headers(club_out).to_excel(
            writer,
            sheet_name="Club_Results",
            index=False,
        )

    # Post-process: apply cutoff borders and auto-width
    wb = load_workbook(output_path)

    for sheet_name, limit in cutoff_info.items():
        if sheet_name in wb.sheetnames and limit > 0:
            ws = wb[sheet_name]
            _apply_cutoff_border(ws, limit)
            _auto_width(ws)

    for sheet_name in (
        "Qualification_Entry_Numbers",
        "Rank_to_Qualify",
        "Score_to_Qualify",
        "Club_Results",
    ):
        if sheet_name in wb.sheetnames:
            _auto_width(wb[sheet_name])

    wb.save(output_path)
    logger.info(f"Output generated: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    params = parse_input_params("inputParams.txt")
    tournament_ids: list = params["TournamentIDs"]
    finals_capacity: int = params["finalsCapacity"]
    club_code: int = params["ClubCode"]

    logger.info("Starting data scrape...")
    df = scrape_all_data(tournament_ids)

    logger.info("Parsing data...")
    parsed = parse_data(df, club_code, finals_capacity)

    logger.info("Writing output...")
    write_output(parsed, "bucs_finals_qualification.xlsx")


if __name__ == "__main__":
    main()
